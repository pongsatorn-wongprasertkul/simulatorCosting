from pathlib import Path
from datetime import datetime
from html import escape
from io import BytesIO
from textwrap import dedent

import numpy as np
from openpyxl import Workbook
import pandas as pd
import plotly.graph_objects as go
import streamlit as st

from app.engine.offset import OFFSET_CATEGORY_TO_COST_ELEMENT, calculate_offset_mitigation
from app.i18n import THAI_LANGUAGE, TRANSLATIONS, translate, translate_value

DATA_DIR = Path(__file__).resolve().parents[2] / "data"
PARTS_PATH = DATA_DIR / "vehicle_parts.csv"
BREAKDOWN_PATH = DATA_DIR / "part_cost_breakdown.csv"
SUPPLIER_PATH = DATA_DIR / "supplier_master.csv"
SUPPLIER_MAP_PATH = DATA_DIR / "part_supplier_map.csv"
SCENARIO_HISTORY_PATH = DATA_DIR / "scenario_history.csv"

VEHICLE_CODE = "BYD-SEAL-EV-TH"
VEHICLE_NAME = "BYD Seal EV Enterprise Cost Model"

DRIVER_COLUMNS = {
    "Oil Price": "oil_factor",
    "FX Rate": "fx_factor",
    "Steel Price": "steel_factor",
    "Copper Price": "copper_factor",
    "Lithium Price": "lithium_factor",
    "Electricity Cost": "electricity_factor",
    "Labor Cost": "labor_factor",
    "Shipping Cost": "shipping_factor",
    "Inflation": "inflation_factor",
}

DRIVER_PRICE_DEFAULTS = {
    "Oil Price": {"unit": "USD/bbl", "current": 82.0, "scenario": 82.0},
    "FX Rate": {"unit": "THB/USD", "current": 35.5, "scenario": 35.5},
    "Steel Price": {"unit": "THB/ton", "current": 24500.0, "scenario": 24500.0},
    "Copper Price": {"unit": "USD/ton", "current": 9200.0, "scenario": 9200.0},
    "Lithium Price": {"unit": "USD/ton LCE", "current": 14500.0, "scenario": 14500.0},
    "Electricity Cost": {"unit": "THB/kWh", "current": 4.25, "scenario": 4.25},
    "Labor Cost": {"unit": "Index", "current": 100.0, "scenario": 100.0},
    "Shipping Cost": {"unit": "USD/container", "current": 1850.0, "scenario": 1850.0},
    "Inflation": {"unit": "Index", "current": 100.0, "scenario": 100.0},
}

COST_ELEMENT_DRIVER_MAP = {
    "Raw Material Cost": ["Steel Price", "Copper Price", "Lithium Price"],
    "Purchased Parts Cost": ["FX Rate", "Steel Price", "Copper Price", "Inflation"],
    "Direct Labor Cost": ["Labor Cost", "Inflation"],
    "Indirect Labor Cost": ["Labor Cost", "Inflation"],
    "Machine Cost": ["Electricity Cost"],
    "Energy Cost": ["Electricity Cost", "Oil Price", "Inflation"],
    "Tooling Cost": ["Steel Price", "FX Rate"],
    "Logistics Cost": ["Shipping Cost", "Oil Price", "FX Rate", "Inflation"],
    "Import Duty": ["FX Rate"],
    "Quality Cost": ["Labor Cost", "FX Rate", "Inflation"],
    "Warranty Reserve": ["Labor Cost", "Lithium Price"],
    "R&D Allocation": ["Labor Cost", "FX Rate", "Inflation"],
    "Software Cost": ["Labor Cost", "FX Rate", "Inflation"],
    "Semiconductor Cost": ["FX Rate", "Copper Price"],
    "Corporate SG&A Allocation": ["Labor Cost", "FX Rate", "Inflation"],
    "Financing Cost": ["FX Rate"],
    "Risk Reserve": ["FX Rate", "Lithium Price", "Shipping Cost", "Inflation"],
    "ESG/Carbon Cost": ["Electricity Cost", "Oil Price"],
}

MONTE_CARLO_RUNS = 10000
SUPPLIER_RISK_SCORES = {
    "Low": 25,
    "Medium": 50,
    "High": 75,
    "Critical": 100,
}

DRIVER_ROOT_CAUSES = {
    "Oil Price": "higher petroleum-linked energy, freight, and carbon-related operating cost",
    "FX Rate": "weaker THB against imported parts, semiconductors, duties, and software cost",
    "Steel Price": "higher body, chassis, tooling, and stamped component material cost",
    "Copper Price": "higher e-motor, wiring, inverter, and electronics material cost",
    "Lithium Price": "higher LFP cell, battery reserve, and battery pack raw material cost",
    "Electricity Cost": "higher plant energy, machine conversion, and ESG/carbon cost",
    "Labor Cost": "higher direct labor, indirect labor, quality, software, and R&D allocation",
    "Shipping Cost": "higher inbound and outbound logistics exposure",
    "Inflation": "broad cost escalation across purchased parts, labor, logistics, SG&A, and reserves",
}

st.set_page_config(page_title="BYD Enterprise Automotive Cost Simulation", layout="wide")

DASHBOARD_CSS = """
    <style>
    .block-container {
        padding-top: 1.4rem;
        max-width: 1480px;
    }
    [data-testid="stMetric"] {
        background: linear-gradient(180deg, #ffffff 0%, #f8fafc 100%);
        border: 1px solid #e2e8f0;
        border-radius: 10px;
        padding: 18px 20px;
        box-shadow: 0 8px 24px rgba(15, 23, 42, 0.07);
        min-height: 118px;
        overflow: visible;
    }
    [data-testid="stMetric"] > div {
        overflow: visible;
        width: 100%;
    }
    [data-testid="stMetricLabel"] {
        color: #475569;
        font-size: 0.82rem;
        font-weight: 650;
        line-height: 1.25;
        min-height: 2.1rem;
        white-space: normal;
        overflow: visible;
        text-overflow: clip;
    }
    [data-testid="stMetricValue"] {
        color: #0f172a;
        font-size: clamp(1.08rem, 1.55vw, 1.75rem);
        font-weight: 760;
        line-height: 1.16;
        white-space: normal;
        word-break: keep-all;
        overflow-wrap: normal;
        overflow: visible;
        text-overflow: clip;
        letter-spacing: 0;
        max-width: 100%;
    }
    [data-testid="stMetricValue"] > div {
        white-space: normal;
        overflow: visible;
        text-overflow: clip;
        max-width: 100%;
    }
    [data-testid="stMetricDelta"] {
        color: #64748b;
        font-size: 0.86rem;
        white-space: normal;
        overflow: visible;
    }
    .exec-hero {
        background: linear-gradient(135deg, #111827 0%, #1f2937 58%, #8f1118 100%);
        color: #ffffff;
        border-radius: 18px;
        padding: 26px 30px;
        margin: 10px 0 22px 0;
        box-shadow: 0 22px 50px rgba(15, 23, 42, 0.24);
    }
    .exec-eyebrow {
        color: #fecaca;
        font-size: 0.78rem;
        font-weight: 760;
        letter-spacing: 0;
        text-transform: uppercase;
        margin-bottom: 8px;
    }
    .exec-title {
        font-size: clamp(1.5rem, 2.2vw, 2.35rem);
        font-weight: 780;
        line-height: 1.12;
        margin-bottom: 8px;
    }
    .exec-subtitle {
        color: #e5e7eb;
        font-size: 0.98rem;
        line-height: 1.5;
        max-width: 980px;
    }
    .exec-kpi-grid {
        display: grid;
        grid-template-columns: repeat(3, minmax(0, 1fr));
        gap: 18px;
        margin: 16px 0 22px 0;
    }
    .exec-kpi-card {
        background: #ffffff;
        border: 1px solid #e5e7eb;
        border-radius: 16px;
        padding: 20px 22px;
        box-shadow: 0 14px 34px rgba(15, 23, 42, 0.10);
        min-height: 154px;
        overflow: visible;
    }
    .exec-kpi-topline {
        display: flex;
        align-items: center;
        justify-content: space-between;
        gap: 12px;
        margin-bottom: 16px;
    }
    .exec-kpi-label {
        color: #475569;
        font-size: 0.82rem;
        font-weight: 760;
        line-height: 1.25;
    }
    .exec-kpi-icon {
        width: 34px;
        height: 34px;
        display: inline-flex;
        align-items: center;
        justify-content: center;
        border-radius: 999px;
        font-weight: 850;
        flex: 0 0 auto;
    }
    .exec-kpi-value {
        color: #0f172a;
        font-size: clamp(1.28rem, 1.85vw, 2.15rem);
        font-weight: 820;
        line-height: 1.08;
        white-space: normal;
        overflow-wrap: anywhere;
        letter-spacing: 0;
    }
    .exec-kpi-delta {
        margin-top: 12px;
        color: #64748b;
        font-size: 0.88rem;
        font-weight: 650;
    }
    .exec-status-banner {
        border-radius: 18px;
        padding: 24px 28px;
        margin: 16px 0 22px 0;
        color: #ffffff;
        box-shadow: 0 18px 44px rgba(15, 23, 42, 0.18);
        display: flex;
        align-items: center;
        justify-content: space-between;
        gap: 18px;
    }
    .exec-status-title {
        font-size: clamp(1.5rem, 2.4vw, 2.4rem);
        font-weight: 880;
        line-height: 1;
        margin-bottom: 8px;
    }
    .exec-status-copy {
        font-size: 0.98rem;
        opacity: 0.94;
        line-height: 1.45;
    }
    .exec-status-pill {
        border: 1px solid rgba(255, 255, 255, 0.42);
        border-radius: 999px;
        padding: 10px 16px;
        font-weight: 800;
        white-space: nowrap;
        background: rgba(255, 255, 255, 0.14);
    }
    .exec-panel {
        background: #ffffff;
        border: 1px solid #e2e8f0;
        border-radius: 16px;
        padding: 22px 24px;
        box-shadow: 0 14px 32px rgba(15, 23, 42, 0.08);
        margin: 14px 0;
    }
    .exec-panel-title {
        color: #0f172a;
        font-size: 1.02rem;
        font-weight: 820;
        margin-bottom: 12px;
    }
    .exec-bullets {
        margin: 0;
        padding-left: 18px;
        color: #334155;
        line-height: 1.55;
        font-size: 0.95rem;
    }
    .exec-risk-grid {
        display: grid;
        grid-template-columns: repeat(3, minmax(0, 1fr));
        gap: 16px;
        margin: 12px 0 18px 0;
    }
    .exec-risk-card {
        background: #ffffff;
        border: 1px solid #e2e8f0;
        border-left: 6px solid #dc2626;
        border-radius: 15px;
        padding: 18px 18px;
        box-shadow: 0 12px 28px rgba(15, 23, 42, 0.08);
    }
    .exec-risk-title {
        color: #0f172a;
        font-size: 0.98rem;
        font-weight: 820;
        margin-bottom: 8px;
    }
    .exec-risk-meta {
        color: #64748b;
        font-size: 0.84rem;
        line-height: 1.45;
    }
    .exec-risk-impact {
        color: #991b1b;
        font-size: 1.08rem;
        font-weight: 800;
        margin-top: 12px;
    }
    .exec-scenario-grid {
        display: grid;
        grid-template-columns: repeat(4, minmax(0, 1fr));
        gap: 14px;
        margin-top: 12px;
    }
    .exec-mini-card {
        background: #f8fafc;
        border: 1px solid #e2e8f0;
        border-radius: 14px;
        padding: 14px 16px;
    }
    .exec-mini-label {
        color: #64748b;
        font-size: 0.78rem;
        font-weight: 760;
        margin-bottom: 6px;
    }
    .exec-mini-value {
        color: #0f172a;
        font-size: 1.02rem;
        font-weight: 820;
        overflow-wrap: anywhere;
    }
    @media (max-width: 1120px) {
        .exec-kpi-grid,
        .exec-risk-grid {
            grid-template-columns: repeat(2, minmax(0, 1fr));
        }
        .exec-scenario-grid {
            grid-template-columns: repeat(2, minmax(0, 1fr));
        }
    }
    @media (max-width: 720px) {
        .exec-kpi-grid,
        .exec-risk-grid,
        .exec-scenario-grid {
            grid-template-columns: 1fr;
        }
        .exec-status-banner {
            align-items: flex-start;
            flex-direction: column;
        }
        .exec-hero {
            padding: 22px 20px;
        }
    }
    @media (max-width: 1100px) {
        [data-testid="stMetric"] {
            padding: 15px 16px;
            min-height: 108px;
        }
        [data-testid="stMetricValue"] {
            font-size: 1.08rem;
        }
    }
    h2, h3 {letter-spacing: 0;}
    </style>
"""

st.markdown(DASHBOARD_CSS, unsafe_allow_html=True)

selected_language = st.sidebar.selectbox(
    translate("English", "language"),
    list(TRANSLATIONS.keys()),
    index=0,
)


def tr(key: str) -> str:
    return translate(selected_language, key)


def tr_value(value: object) -> object:
    return translate_value(selected_language, value)


def tr_root_cause(driver_name: str) -> str:
    return tr(f"root_cause_{driver_name}")


def render_html(html: str) -> None:
    st.markdown(dedent(html).strip(), unsafe_allow_html=True)


dashboard_view = st.sidebar.radio(
    tr("dashboard_view"),
    ["executive", "analyst"],
    format_func=lambda value: tr("executive_dashboard")
    if value == "executive"
    else tr("analyst_dashboard"),
)


st.title(tr("app_title"))
st.caption(f"{VEHICLE_CODE} | {VEHICLE_NAME}")


@st.cache_data
def load_vehicle_data() -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    return (
        pd.read_csv(PARTS_PATH),
        pd.read_csv(BREAKDOWN_PATH),
        pd.read_csv(SUPPLIER_PATH),
        pd.read_csv(SUPPLIER_MAP_PATH),
    )


def format_thb(value: float) -> str:
    if selected_language == THAI_LANGUAGE:
        return f"฿{value:,.0f}"
    return f"THB {value:,.0f}"


def format_pct(value: float) -> str:
    return f"{value:.2f}%"


def render_exec_kpi(
    label: str,
    value: str,
    delta: str,
    icon: str,
    color: str,
    trend: str,
) -> str:
    soft_color = f"{color}18"
    return dedent(f"""
    <div class="exec-kpi-card">
        <div class="exec-kpi-topline">
            <div class="exec-kpi-label">{escape(label)}</div>
            <div class="exec-kpi-icon" style="background:{soft_color}; color:{color};">{icon}</div>
        </div>
        <div class="exec-kpi-value">{escape(value)}</div>
        <div class="exec-kpi-delta" style="color:{color};">{trend} {escape(delta)}</div>
    </div>
    """).strip()


def render_exec_risk_card(title: str, meta: str, impact: str, color: str) -> str:
    return dedent(f"""
    <div class="exec-risk-card" style="border-left-color:{color};">
        <div class="exec-risk-title">{escape(title)}</div>
        <div class="exec-risk-meta">{escape(meta)}</div>
        <div class="exec-risk-impact" style="color:{color};">{escape(impact)}</div>
    </div>
    """).strip()


def build_gauge_chart(
    title: str,
    value: float,
    target: float,
    max_value: float,
    color: str,
) -> go.Figure:
    figure = go.Figure(
        go.Indicator(
            mode="gauge+number+delta",
            value=value,
            number={"suffix": "%", "font": {"size": 34, "color": "#0f172a"}},
            delta={"reference": target, "suffix": "%", "font": {"size": 14}},
            title={"text": title, "font": {"size": 15, "color": "#334155"}},
            gauge={
                "axis": {"range": [0, max_value], "tickcolor": "#94a3b8"},
                "bar": {"color": color, "thickness": 0.28},
                "bgcolor": "#f8fafc",
                "borderwidth": 0,
                "steps": [
                    {"range": [0, target], "color": "#fee2e2"},
                    {"range": [target, min(max_value, target + 6)], "color": "#fef3c7"},
                    {"range": [min(max_value, target + 6), max_value], "color": "#dcfce7"},
                ],
                "threshold": {
                    "line": {"color": "#111827", "width": 3},
                    "thickness": 0.78,
                    "value": target,
                },
            },
        )
    )
    figure.update_layout(
        height=300,
        margin={"l": 18, "r": 18, "t": 50, "b": 18},
        paper_bgcolor="white",
        font={"family": "Arial", "color": "#0f172a"},
    )
    return figure


def localize_dataframe(frame: pd.DataFrame) -> pd.DataFrame:
    localized = frame.copy()
    for column in localized.select_dtypes(include="object").columns:
        localized[column] = localized[column].map(tr_value)
    return localized


def load_scenario_history() -> pd.DataFrame:
    if not SCENARIO_HISTORY_PATH.exists():
        return pd.DataFrame()
    return pd.read_csv(SCENARIO_HISTORY_PATH)


def save_scenario(row: dict[str, object]) -> None:
    history = load_scenario_history()
    updated_history = pd.concat([history, pd.DataFrame([row])], ignore_index=True)
    updated_history.to_csv(SCENARIO_HISTORY_PATH, index=False)


def reset_to_base_case() -> None:
    for driver_name, defaults in DRIVER_PRICE_DEFAULTS.items():
        current_value = float(defaults["current"])
        st.session_state[f"{driver_name}_current"] = current_value
        st.session_state[f"{driver_name}_scenario"] = current_value
        st.session_state[f"{driver_name}_mc_min"] = current_value
        st.session_state[f"{driver_name}_mc_max"] = current_value
        st.session_state[f"{driver_name}_mc_mean"] = current_value
        st.session_state[f"{driver_name}_mc_volatility"] = 0.0

    st.session_state["supplier_price_increase"] = 0
    st.session_state["supplier_delay_days"] = 0
    st.session_state["supplier_disruption_rate"] = 0
    st.session_state["delay_penalty_rate"] = 0.0

    for category in OFFSET_CATEGORY_TO_COST_ELEMENT:
        st.session_state[f"offset_{category}"] = 0.0


def scenario_history_to_excel(history: pd.DataFrame) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Scenario History"

    for column_index, column_name in enumerate(history.columns, start=1):
        worksheet.cell(row=1, column=column_index, value=column_name)

    for row_index, (_, row) in enumerate(history.iterrows(), start=2):
        for column_index, value in enumerate(row.tolist(), start=1):
            worksheet.cell(row=row_index, column=column_index, value=value)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def calculate_change_percent(current_price: float, scenario_price: float) -> float:
    if current_price == 0:
        return 0.0
    return (scenario_price - current_price) / current_price


def weighted_cost_element_impact(
    row: pd.Series,
    cost_element: str,
    driver_changes: dict[str, float],
) -> float:
    drivers = COST_ELEMENT_DRIVER_MAP[cost_element]
    sensitivity = sum(row[DRIVER_COLUMNS[driver]] for driver in drivers) / len(drivers)
    change = sum(driver_changes[driver] for driver in drivers) / len(drivers)
    return row["cost_amount"] * sensitivity * change


def calculate_element_impacts(
    breakdown: pd.DataFrame,
    changes: dict[str, float],
) -> pd.DataFrame:
    impacted = breakdown.copy()
    impacted["element_impact"] = impacted.apply(
        lambda row: weighted_cost_element_impact(row, row["cost_element"], changes),
        axis=1,
    )
    return impacted


def classify_risk(score: float) -> str:
    if score >= 75:
        return "Critical"
    if score >= 50:
        return "High"
    if score >= 25:
        return "Medium"
    return "Low"


def classify_monte_carlo_risk(probability: float) -> str:
    if probability >= 0.50:
        return "Critical Risk"
    if probability >= 0.30:
        return "High Risk"
    if probability >= 0.15:
        return "Medium Risk"
    return "Low Risk"


def build_row_driver_weights(breakdown: pd.DataFrame) -> np.ndarray:
    weights = np.zeros((len(breakdown), len(DRIVER_COLUMNS)))
    driver_names = list(DRIVER_COLUMNS)
    for row_index, (_, row) in enumerate(breakdown.iterrows()):
        mapped_drivers = COST_ELEMENT_DRIVER_MAP[row["cost_element"]]
        sensitivity = (
            sum(row[DRIVER_COLUMNS[driver]] for driver in mapped_drivers)
            / len(mapped_drivers)
        )
        for driver in mapped_drivers:
            weights[row_index, driver_names.index(driver)] = sensitivity / len(mapped_drivers)
    return weights


def explain_cost_increase(row: pd.Series) -> str:
    direction = "increased" if row["driver_impact"] >= 0 else "decreased"
    return (
        f"{row['Driver']} {direction} {row['part_name']} in {row['module_name']} "
        f"through {row['cost_element']}: {tr_root_cause(row['Driver'])}."
    )


def validate_inputs(
    parts: pd.DataFrame,
    breakdown: pd.DataFrame,
    scenarios: dict[str, dict[str, float]],
) -> list[str]:
    errors = []

    for driver_name, scenario in scenarios.items():
        if scenario["current_price"] <= 0:
            errors.append(f"{driver_name}: current price must be greater than 0.")
        if scenario["scenario_price"] < 0:
            errors.append(f"{driver_name}: scenario price must be greater than or equal to 0.")

    if (parts["sales_price"] <= 0).any():
        bad_parts = parts.loc[parts["sales_price"] <= 0, "part_name"].tolist()
        errors.append(
            "Sales price must be greater than 0 for every part. "
            f"Please check: {', '.join(bad_parts[:5])}."
        )

    if (parts["total_base_cost"] < 0).any():
        bad_parts = parts.loc[parts["total_base_cost"] < 0, "part_name"].tolist()
        errors.append(
            "Cost values must not be negative. "
            f"Please check base cost for: {', '.join(bad_parts[:5])}."
        )

    if (breakdown["cost_amount"] < 0).any():
        bad_codes = breakdown.loc[breakdown["cost_amount"] < 0, "part_code"].unique().tolist()
        errors.append(
            "Cost breakdown values must not be negative. "
            f"Please check part codes: {', '.join(bad_codes[:5])}."
        )

    return errors


parts_df, breakdown_df, suppliers_df, supplier_map_df = load_vehicle_data()
if "inflation_factor" not in parts_df.columns:
    parts_df["inflation_factor"] = 0.10

ALL_MODULES = "All modules"
module_options = [ALL_MODULES] + sorted(parts_df["module_name"].unique().tolist())

with st.sidebar:
    if st.button(tr("reset_to_base_case"), use_container_width=True):
        reset_to_base_case()

    st.subheader(tr("scenario_management"))
    scenario_name = st.text_input(tr("scenario_name"), "Base commodity scenario")

    st.subheader(tr("commodity_scenario_inputs"))
    price_scenarios = {}
    driver_changes = {}
    for driver_name, defaults in DRIVER_PRICE_DEFAULTS.items():
        st.markdown(f"**{tr(driver_name)}**")
        current_col, scenario_col = st.columns(2)
        current_price = current_col.number_input(
            tr("current"),
            min_value=0.0001,
            value=float(defaults["current"]),
            step=max(float(defaults["current"]) * 0.01, 0.1),
            key=f"{driver_name}_current",
            help=str(defaults["unit"]),
        )
        scenario_price = scenario_col.number_input(
            tr("scenario"),
            min_value=0.0,
            value=float(defaults["scenario"]),
            step=max(float(defaults["current"]) * 0.01, 0.1),
            key=f"{driver_name}_scenario",
            help=str(defaults["unit"]),
        )
        change_percent = calculate_change_percent(current_price, scenario_price)
        st.caption(f"{defaults['unit']} | {tr('impact')} {change_percent:.2%}")
        price_scenarios[driver_name] = {
            "unit": defaults["unit"],
            "current_price": current_price,
            "scenario_price": scenario_price,
            "change_percent": change_percent,
        }
        driver_changes[driver_name] = change_percent

    st.subheader(tr("monte_carlo_assumptions"))
    gp_target_percent = st.number_input(
        tr("target_gp"),
        min_value=0.0,
        max_value=100.0,
        value=12.0,
        step=0.5,
    )
    monte_carlo_config = {}
    for driver_name, defaults in DRIVER_PRICE_DEFAULTS.items():
        with st.expander(driver_name, expanded=False):
            current_value = float(price_scenarios[driver_name]["current_price"])
            scenario_value = float(price_scenarios[driver_name]["scenario_price"])
            lower_default = scenario_value
            upper_default = scenario_value
            mean_default = scenario_value
            volatility_default = 0.0
            min_value = st.number_input(
                tr("min_value"),
                min_value=0.0001,
                value=float(lower_default),
                step=max(current_value * 0.01, 0.1),
                key=f"{driver_name}_mc_min",
            )
            max_value = st.number_input(
                tr("max_value"),
                min_value=0.0001,
                value=float(upper_default),
                step=max(current_value * 0.01, 0.1),
                key=f"{driver_name}_mc_max",
            )
            mean_value = st.number_input(
                tr("mean_value"),
                min_value=0.0001,
                value=float(mean_default),
                step=max(current_value * 0.01, 0.1),
                key=f"{driver_name}_mc_mean",
            )
            volatility = st.number_input(
                tr("volatility"),
                min_value=0.0,
                value=float(volatility_default),
                step=max(current_value * 0.005, 0.1),
                key=f"{driver_name}_mc_volatility",
            )
            monte_carlo_config[driver_name] = {
                "min": min(min_value, max_value),
                "max": max(min_value, max_value),
                "mean": mean_value,
                "volatility": volatility,
            }

    st.subheader(tr("supplier_scenario"))
    selected_supplier = st.selectbox(
        tr("disrupted_supplier"),
        [tr("no_disruption")] + sorted(suppliers_df["supplier_name"].tolist()),
    )
    supplier_price_increase = (
        st.slider(tr("supplier_price_increase"), 0, 30, 0, key="supplier_price_increase")
        / 100
    )
    supplier_delay_days = st.number_input(
        tr("supplier_delay_days"),
        min_value=0,
        max_value=120,
        value=0,
        step=1,
        key="supplier_delay_days",
    )
    supplier_disruption_rate = (
        st.slider(
            tr("supplier_disruption_severity"),
            0,
            100,
            0,
            key="supplier_disruption_rate",
        )
        / 100
    )
    delay_penalty_rate = st.number_input(
        tr("delay_penalty_per_day"),
        min_value=0.0,
        max_value=5.0,
        value=0.0,
        step=0.01,
        key="delay_penalty_rate",
    ) / 100

    st.subheader(tr("cost_offset_mitigation"))
    offset_reductions = {}
    for category in OFFSET_CATEGORY_TO_COST_ELEMENT:
        offset_reductions[category] = (
            st.number_input(
                tr(category),
                min_value=0.0,
                max_value=100.0,
                value=0.0,
                step=0.5,
                key=f"offset_{category}",
            )
            / 100
        )

    st.subheader(tr("bom_drilldown"))
    selected_module = st.selectbox(
        tr("select_module"),
        module_options,
        format_func=lambda value: tr_value(value),
    )
    part_scope = parts_df
    if selected_module != ALL_MODULES:
        part_scope = parts_df[parts_df["module_name"] == selected_module]
    selected_part_name = st.selectbox(tr("select_part"), part_scope["part_name"].tolist())

validation_errors = validate_inputs(parts_df, breakdown_df, price_scenarios)
if validation_errors:
    for error in validation_errors:
        st.error(error)
    st.stop()

breakdown_enriched = breakdown_df.merge(parts_df, on="part_code", how="left")
breakdown_enriched = calculate_element_impacts(breakdown_enriched, driver_changes)

traceability_frames = []
has_driver_changes = any(abs(change) > 0.000001 for change in driver_changes.values())
for trace_driver in DRIVER_COLUMNS:
    trace_changes = {name: 0.0 for name in DRIVER_COLUMNS}
    trace_changes[trace_driver] = driver_changes[trace_driver]
    trace_frame = calculate_element_impacts(breakdown_enriched, trace_changes)
    if has_driver_changes:
        trace_frame = trace_frame[trace_frame["element_impact"].abs() > 0].copy()
    else:
        trace_frame = trace_frame.head(1).copy()
    trace_frame["Driver"] = trace_driver
    trace_frame["driver_impact"] = trace_frame["element_impact"]
    trace_frame["Root cause"] = tr_root_cause(trace_driver)
    traceability_frames.append(trace_frame)

traceability_df = pd.concat(traceability_frames, ignore_index=True)
traceability_df["Explanation"] = traceability_df.apply(explain_cost_increase, axis=1)

part_impacts = (
    breakdown_enriched.groupby("part_code", as_index=False)
    .agg(
        old_cost=("cost_amount", "sum"),
        impact_amount=("element_impact", "sum"),
    )
)
results_df = parts_df.merge(part_impacts, on="part_code", how="left")
results_df["new_cost"] = results_df["old_cost"] + results_df["impact_amount"]
results_df["gp_amount"] = results_df["sales_price"] - results_df["new_cost"]
results_df["gp_percent"] = results_df["gp_amount"] / results_df["sales_price"] * 100
results_df["risk_exposure"] = (
    results_df["risk_score"] * results_df["impact_amount"].abs() / 100
)
risk_cutoff = results_df["risk_exposure"].quantile(0.75)
results_df["sensitivity_score"] = (
    results_df["impact_amount"].abs() / results_df["old_cost"].replace(0, pd.NA) * 100
).fillna(0)
results_df["risk_band"] = results_df["risk_exposure"].apply(
    lambda value: classify_risk((value / risk_cutoff * 50) if risk_cutoff else 0)
)

total_sales_price = results_df["sales_price"].sum()
total_base_cost = results_df["old_cost"].sum()
total_new_cost = results_df["new_cost"].sum()
total_impact = total_new_cost - total_base_cost
total_gp = total_sales_price - total_new_cost
total_gp_percent = total_gp / total_sales_price * 100 if total_sales_price else 0

supplier_detail_df = (
    supplier_map_df.merge(results_df, on="part_code", how="left")
    .merge(suppliers_df, on="supplier_name", how="left")
)
supplier_detail_df["supplier_cost_base"] = (
    supplier_detail_df["new_cost"] * supplier_detail_df["allocation_percent"]
)
supplier_detail_df["fx_exposure"] = np.where(
    supplier_detail_df["currency"].eq("THB"),
    0.0,
    driver_changes["FX Rate"],
)
supplier_detail_df["supplier_price_impact"] = (
    supplier_detail_df["supplier_cost_base"] * supplier_price_increase
)
supplier_detail_df["supplier_delay_impact"] = (
    supplier_detail_df["supplier_cost_base"] * supplier_delay_days * delay_penalty_rate
)
supplier_detail_df["supplier_disruption_impact"] = np.where(
    supplier_detail_df["supplier_name"].eq(selected_supplier),
    supplier_detail_df["supplier_cost_base"] * supplier_disruption_rate,
    0.0,
)
supplier_detail_df["supplier_fx_impact"] = (
    supplier_detail_df["supplier_cost_base"] * supplier_detail_df["fx_exposure"]
)
supplier_detail_df["supplier_total_impact"] = (
    supplier_detail_df["supplier_price_impact"]
    + supplier_detail_df["supplier_delay_impact"]
    + supplier_detail_df["supplier_disruption_impact"]
    + supplier_detail_df["supplier_fx_impact"]
)
supplier_detail_df["risk_score"] = supplier_detail_df["risk_level"].map(SUPPLIER_RISK_SCORES)

supplier_summary_df = (
    supplier_detail_df.groupby(
        ["supplier_name", "country", "currency", "risk_level"],
        as_index=False,
    )
    .agg(
        supplied_parts=("part_code", "nunique"),
        supplier_cost_base=("supplier_cost_base", "sum"),
        supplier_price_impact=("supplier_price_impact", "sum"),
        supplier_delay_impact=("supplier_delay_impact", "sum"),
        supplier_disruption_impact=("supplier_disruption_impact", "sum"),
        supplier_fx_impact=("supplier_fx_impact", "sum"),
        supplier_total_impact=("supplier_total_impact", "sum"),
        risk_score=("risk_score", "max"),
    )
)
supplier_summary_df["vehicle_cost_contribution %"] = (
    supplier_summary_df["supplier_cost_base"] / total_new_cost * 100
    if total_new_cost
    else 0
)
supplier_summary_df["risk_exposure"] = (
    supplier_summary_df["risk_score"] * supplier_summary_df["supplier_cost_base"] / 100
    + supplier_summary_df["supplier_total_impact"].abs()
)

is_base_case = (
    abs(total_impact) < 0.01
    and all(abs(change) < 0.000001 for change in driver_changes.values())
    and supplier_price_increase == 0
    and supplier_delay_days == 0
    and supplier_disruption_rate == 0
    and delay_penalty_rate == 0
    and all(reduction == 0 for reduction in offset_reductions.values())
)

filtered_df = results_df
if selected_module != ALL_MODULES:
    filtered_df = results_df[results_df["module_name"] == selected_module]

selected_part = results_df[results_df["part_name"] == selected_part_name].iloc[0]
selected_breakdown = breakdown_enriched[
    breakdown_enriched["part_code"] == selected_part["part_code"]
].copy()
selected_breakdown["new_cost_amount"] = (
    selected_breakdown["cost_amount"] + selected_breakdown["element_impact"]
)

if dashboard_view == "executive":
    base_gp = total_sales_price - total_base_cost
    base_gp_percent = base_gp / total_sales_price * 100 if total_sales_price else 0
    cost_change_percent = total_impact / total_base_cost * 100 if total_base_cost else 0

    executive_driver_rows = []
    for driver_name, factor_column in DRIVER_COLUMNS.items():
        executive_driver_rows.append(
            {
                "Driver": driver_name,
                "Impact": (
                    parts_df["total_base_cost"]
                    * parts_df[factor_column]
                    * driver_changes[driver_name]
                ).sum(),
                "Change %": driver_changes[driver_name] * 100,
            }
        )
    executive_driver_df = pd.DataFrame(executive_driver_rows).sort_values(
        "Impact",
        key=lambda values: values.abs(),
        ascending=False,
    )

    quick_rng = np.random.default_rng(42)
    quick_driver_changes = []
    quick_driver_names = list(DRIVER_COLUMNS)
    for driver_name in quick_driver_names:
        config = monte_carlo_config[driver_name]
        current_price = price_scenarios[driver_name]["current_price"]
        sampled_values = quick_rng.normal(
            loc=config["mean"],
            scale=config["volatility"],
            size=MONTE_CARLO_RUNS,
        )
        sampled_values = np.clip(sampled_values, config["min"], config["max"])
        quick_driver_changes.append((sampled_values - current_price) / current_price)
    quick_driver_matrix = np.column_stack(quick_driver_changes)
    quick_row_weights = build_row_driver_weights(breakdown_enriched)
    quick_row_change = quick_driver_matrix @ quick_row_weights.T
    quick_cost_impact = quick_row_change * breakdown_enriched["cost_amount"].to_numpy()
    quick_total_cost = total_base_cost + quick_cost_impact.sum(axis=1)
    quick_gp_percent = (total_sales_price - quick_total_cost) / total_sales_price * 100
    worst_case_gp_percent = float(np.percentile(quick_gp_percent, 5))

    supplier_risk_level = supplier_summary_df.sort_values(
        "risk_exposure",
        ascending=False,
    ).iloc[0]["risk_level"]

    if is_base_case:
        traffic_status = "safe"
        traffic_color = "#16a34a"
    elif total_gp_percent < gp_target_percent or worst_case_gp_percent < gp_target_percent:
        traffic_status = "critical"
        traffic_color = "#dc2626"
    elif total_gp_percent < gp_target_percent + 3 or abs(cost_change_percent) > 5:
        traffic_status = "warning"
        traffic_color = "#ca8a04"
    else:
        traffic_status = "safe"
        traffic_color = "#16a34a"
    status_icon = {"safe": "&#10003;", "warning": "!", "critical": "&#10005;"}[traffic_status]
    status_copy = {
        "safe": tr("status_safe_message"),
        "warning": tr("status_warning_message"),
        "critical": tr("status_critical_message"),
    }[traffic_status]
    status_gradient = {
        "safe": "linear-gradient(135deg, #047857 0%, #16a34a 100%)",
        "warning": "linear-gradient(135deg, #a16207 0%, #eab308 100%)",
        "critical": "linear-gradient(135deg, #991b1b 0%, #dc2626 100%)",
    }[traffic_status]

    top_drivers = executive_driver_df.head(3)
    top_driver_lines = [
        f"{tr(row['Driver'])} {row['Change %']:+.1f}%"
        for _, row in top_drivers.iterrows()
    ]
    gp_sentence_key = (
        "gp_decreased_from_to"
        if total_gp_percent < base_gp_percent
        else "gp_increased_from_to"
    )
    if is_base_case:
        executive_summary = tr("base_case_no_market_impact")
    else:
        executive_summary = (
            tr("vehicle_cost_changed_by").format(
                direction=tr("increased" if total_impact >= 0 else "decreased"),
                percent=format_pct(abs(cost_change_percent)),
            )
            + "\n\n"
            + f"{tr('main_drivers')}:\n"
            + "\n".join(top_driver_lines)
            + "\n\n"
            + tr(gp_sentence_key).format(
                before=format_pct(base_gp_percent),
                after=format_pct(total_gp_percent),
            )
        )

    history_for_exec = load_scenario_history()
    if history_for_exec.empty:
        scenario_compare_df = pd.DataFrame()
    else:
        scenario_compare_df = history_for_exec.sort_values("saved_at", ascending=False).head(5)

    top_module_summary = (
        results_df.groupby("module_name", as_index=False)
        .agg(impact_amount=("impact_amount", "sum"))
        .sort_values("impact_amount", key=lambda values: values.abs(), ascending=False)
    )
    top_part_summary = results_df.sort_values(
        "impact_amount",
        key=lambda values: values.abs(),
        ascending=False,
    )
    top_module_name = (
        tr(top_module_summary.iloc[0]["module_name"])
        if not top_module_summary.empty
        else tr("was_unchanged")
    )
    top_part_name = (
        str(top_part_summary.iloc[0]["part_name"])
        if not top_part_summary.empty
        else tr("was_unchanged")
    )

    render_html(
        f"""
        <div class="exec-hero">
            <div class="exec-eyebrow">{escape(tr('app_title'))}</div>
            <div class="exec-title">{escape(tr('executive_dashboard'))}</div>
            <div class="exec-subtitle">{escape(tr('executive_dashboard_subtitle'))}</div>
        </div>
        """
    )

    cost_trend = "&#8593;" if total_impact > 0 else "&#8595;" if total_impact < 0 else "&#8594;"
    gp_trend = "&#8593;" if total_gp_percent >= base_gp_percent else "&#8595;"
    risk_trend = "&#9679;"
    supplier_trend = "&#9650;" if supplier_summary_df["supplier_total_impact"].abs().sum() > 0 else "&#8594;"
    worst_case_trend = "&#8593;" if worst_case_gp_percent >= gp_target_percent else "&#8595;"
    net_color = "#dc2626" if total_impact > 0 else "#16a34a" if total_impact < 0 else "#64748b"
    gp_color = "#16a34a" if total_gp_percent >= gp_target_percent else "#dc2626"
    risk_color = traffic_color

    kpi_cards = [
        render_exec_kpi(
            tr("total_vehicle_cost"),
            format_thb(total_new_cost),
            tr("cost_delta").format(delta=format_thb(total_impact)),
            "&#3647;",
            net_color,
            cost_trend,
        ),
        render_exec_kpi(
            tr("gp_percent"),
            format_pct(total_gp_percent),
            tr("gp_delta").format(delta=format_pct(total_gp_percent - base_gp_percent)),
            "%",
            gp_color,
            gp_trend,
        ),
        render_exec_kpi(
            tr("net_cost_impact"),
            format_thb(total_impact),
            tr("cost_change_percent").format(delta=format_pct(cost_change_percent)),
            "&#916;",
            net_color,
            cost_trend,
        ),
        render_exec_kpi(
            tr("risk_level"),
            tr(traffic_status).upper(),
            tr("traffic_light_status"),
            status_icon,
            risk_color,
            risk_trend,
        ),
        render_exec_kpi(
            tr("supplier_risk"),
            tr(str(supplier_risk_level)),
            tr("supplier_exposure").format(
                impact=format_thb(supplier_summary_df["supplier_total_impact"].sum())
            ),
            "S",
            "#7c3aed",
            supplier_trend,
        ),
        render_exec_kpi(
            tr("worst_case_gp"),
            format_pct(worst_case_gp_percent),
            tr("target_gp_delta").format(delta=format_pct(worst_case_gp_percent - gp_target_percent)),
            "P5",
            "#0f766e" if worst_case_gp_percent >= gp_target_percent else "#dc2626",
            worst_case_trend,
        ),
    ]
    render_html(f"""<div class="exec-kpi-grid">{''.join(kpi_cards)}</div>""")

    render_html(
        f"""
        <div class="exec-status-banner" style="background:{status_gradient};">
            <div>
                <div class="exec-status-title">{status_icon} {escape(tr(traffic_status).upper())}</div>
                <div class="exec-status-copy">{escape(status_copy)}</div>
            </div>
            <div class="exec-status-pill">{escape(tr('traffic_light_status'))}</div>
        </div>
        """
    )

    if is_base_case:
        summary_bullets = [
            tr("base_case_no_market_impact"),
            tr("current_gp_vs_target").format(
                gp=format_pct(total_gp_percent),
                target=format_pct(gp_target_percent),
            ),
            tr("no_offset_required"),
        ]
    else:
        summary_bullets = [
            tr("vehicle_cost_changed_by").format(
                direction=tr("increased" if total_impact >= 0 else "decreased"),
                percent=format_pct(abs(cost_change_percent)),
            ),
            tr(gp_sentence_key).format(
                before=format_pct(base_gp_percent),
                after=format_pct(total_gp_percent),
            ),
            tr("primary_module_part_summary").format(
                module=top_module_name,
                part=top_part_name,
            ),
        ]
    bullet_html = "".join(f"<li>{escape(item)}</li>" for item in summary_bullets)
    render_html(
        f"""
        <div class="exec-panel">
            <div class="exec-panel-title">{escape(tr('executive_explanation_panel'))}</div>
            <ul class="exec-bullets">{bullet_html}</ul>
        </div>
        """
    )

    risk_cards = []
    risk_palette = ["#dc2626", "#f97316", "#eab308"]
    for risk_index, (_, row) in enumerate(top_drivers.iterrows()):
        impact_value = float(row["Impact"])
        risk_cards.append(
            render_exec_risk_card(
                tr(row["Driver"]),
                tr("risk_card_meta").format(change=format_pct(row["Change %"])),
                format_thb(impact_value),
                risk_palette[min(risk_index, len(risk_palette) - 1)],
            )
        )
    render_html(
        f"""
        <div class="exec-panel">
            <div class="exec-panel-title">{escape(tr('top_risks'))}</div>
            <div class="exec-risk-grid">{''.join(risk_cards)}</div>
        </div>
        """
    )

    recommendation_items = [
        tr("recommendation_ok") if traffic_status == "safe" else tr("recommendation_gap"),
        tr("reduce_labor_cost"),
        tr("negotiate_supplier_discount"),
        tr("increase_selling_price"),
    ]
    recommendation_html = "".join(f"<li>{escape(item)}</li>" for item in recommendation_items)
    render_html(
        f"""
        <div class="exec-panel">
            <div class="exec-panel-title">{escape(tr('recommendation'))}</div>
            <ul class="exec-bullets">{recommendation_html}</ul>
        </div>
        """
    )

    render_html(
        f"""
        <div class="exec-panel">
            <div class="exec-panel-title">{escape(tr('executive_scenario_status'))}</div>
            <div class="exec-scenario-grid">
                <div class="exec-mini-card">
                    <div class="exec-mini-label">{escape(tr('scenario_name'))}</div>
                    <div class="exec-mini-value">{escape(scenario_name)}</div>
                </div>
                <div class="exec-mini-card">
                    <div class="exec-mini-label">{escape(tr('main_drivers'))}</div>
                    <div class="exec-mini-value">{escape(', '.join(top_driver_lines[:3]))}</div>
                </div>
                <div class="exec-mini-card">
                    <div class="exec-mini-label">{escape(tr('target_gp'))}</div>
                    <div class="exec-mini-value">{escape(format_pct(gp_target_percent))}</div>
                </div>
                <div class="exec-mini-card">
                    <div class="exec-mini-label">{escape(tr('scenario_compare'))}</div>
                    <div class="exec-mini-value">{escape(str(len(history_for_exec)))} {escape(tr('saved_scenarios'))}</div>
                </div>
            </div>
        </div>
        """
    )

    chart_cols = st.columns(2, gap="large")
    with chart_cols[0]:
        st.plotly_chart(
            build_gauge_chart(
                tr("gp_gauge"),
                total_gp_percent,
                gp_target_percent,
                max(40.0, gp_target_percent + 20.0, total_gp_percent + 10.0),
                gp_color,
            ),
            use_container_width=True,
        )

    with chart_cols[1]:
        risk_score_value = {"safe": 20.0, "warning": 62.0, "critical": 92.0}[traffic_status]
        st.plotly_chart(
            build_gauge_chart(
                tr("risk_gauge"),
                risk_score_value,
                60.0,
                100.0,
                risk_color,
            ),
            use_container_width=True,
        )

    chart_cols = st.columns(2)
    with chart_cols[0]:
        waterfall_fig = go.Figure(
            go.Waterfall(
                name=tr("cost_impact_waterfall"),
                orientation="v",
                measure=["absolute", "relative", "total"],
                x=[tr("base_cost"), tr("net_cost_impact"), tr("total_vehicle_cost")],
                y=[total_base_cost, total_impact, total_new_cost],
                connector={"line": {"color": "#94a3b8"}},
                increasing={"marker": {"color": "#dc2626"}},
                decreasing={"marker": {"color": "#16a34a"}},
                totals={"marker": {"color": "#1f2937"}},
            )
        )
        waterfall_fig.update_layout(
            title=tr("cost_impact_waterfall"),
            height=360,
            margin={"l": 20, "r": 20, "t": 54, "b": 34},
            paper_bgcolor="white",
            plot_bgcolor="white",
            font={"family": "Arial", "color": "#0f172a"},
            yaxis_title="THB",
        )
        st.plotly_chart(waterfall_fig, use_container_width=True)

    with chart_cols[1]:
        if history_for_exec.empty:
            st.info(tr("no_scenarios_to_compare"))
        else:
            scenario_fig = go.Figure()
            scenario_fig.add_bar(
                x=scenario_compare_df["scenario_name"],
                y=scenario_compare_df["total_vehicle_cost"],
                name=tr("total_vehicle_cost"),
                marker_color="#8f1118",
            )
            scenario_fig.add_scatter(
                x=scenario_compare_df["scenario_name"],
                y=scenario_compare_df["gp_percent"],
                name=tr("gp_percent"),
                yaxis="y2",
                mode="lines+markers",
                line={"color": "#0f766e", "width": 3},
            )
            scenario_fig.update_layout(
                title=tr("scenario_compare"),
                height=360,
                margin={"l": 20, "r": 20, "t": 54, "b": 70},
                paper_bgcolor="white",
                plot_bgcolor="white",
                font={"family": "Arial", "color": "#0f172a"},
                yaxis={"title": "THB"},
                yaxis2={
                    "title": tr("gp_percent"),
                    "overlaying": "y",
                    "side": "right",
                    "ticksuffix": "%",
                },
                legend={"orientation": "h", "y": -0.24},
            )
            st.plotly_chart(scenario_fig, use_container_width=True)

    if not scenario_compare_df.empty:
        st.subheader(tr("scenario_compare"))
        st.dataframe(
            scenario_compare_df[
                ["saved_at", "scenario_name", "total_vehicle_cost", "gp_amount", "gp_percent"]
            ],
            use_container_width=True,
            hide_index=True,
            column_config={
                "total_vehicle_cost": st.column_config.NumberColumn(format="THB %.0f"),
                "gp_amount": st.column_config.NumberColumn(format="THB %.0f"),
                "gp_percent": st.column_config.NumberColumn(format="%.2f%%"),
            },
        )

    st.stop()

kpi_top_cols = st.columns(3, gap="large")
kpi_top_cols[0].metric(tr("vehicle_sales_price"), format_thb(total_sales_price))
kpi_top_cols[1].metric(tr("base_cost"), format_thb(total_base_cost))
kpi_top_cols[2].metric(tr("simulated_cost"), format_thb(total_new_cost), format_thb(total_impact))

kpi_bottom_cols = st.columns(2, gap="large")
kpi_bottom_cols[0].metric(tr("gross_profit"), format_thb(total_gp))
kpi_bottom_cols[1].metric(tr("gp_percent"), format_pct(total_gp_percent))

st.subheader(tr("cost_offset_mitigation"))
offset_result = calculate_offset_mitigation(
    cost_breakdown=breakdown_enriched[["cost_element", "cost_amount"]],
    reduction_percentages=offset_reductions,
    commodity_cost_increase=max(float(total_impact), 0.0),
    base_total_cost=float(total_base_cost),
    sales_price=float(total_sales_price),
    target_gp_percent=float(gp_target_percent),
)
recommendation_message = (
    tr("recommendation_ok")
    if offset_result["target_maintained"]
    else tr("recommendation_gap")
)
target_status = (
    tr("target_maintained")
    if offset_result["target_maintained"]
    else tr("target_not_maintained")
)

offset_cols = st.columns(4, gap="large")
offset_cols[0].metric(
    tr("commodity_cost_increase"),
    format_thb(offset_result["commodity_cost_increase"]),
)
offset_cols[1].metric(tr("offset_savings"), format_thb(offset_result["offset_savings"]))
offset_cols[2].metric(tr("net_cost_impact"), format_thb(offset_result["net_impact"]))
offset_cols[3].metric(tr("required_additional_savings"), format_thb(offset_result["required_additional_savings"]))

gp_offset_top_cols = st.columns(3, gap="large")
gp_offset_top_cols[0].metric(tr("gp_before"), format_thb(offset_result["gp_before"]))
gp_offset_top_cols[1].metric(tr("gp_after"), format_thb(offset_result["gp_after"]))
gp_offset_top_cols[2].metric(tr("target_gp"), format_pct(offset_result["target_gp_percent"]))

gp_offset_bottom_cols = st.columns(2, gap="large")
gp_offset_bottom_cols[0].metric(tr("gap_to_target"), format_pct(offset_result["gap_to_target"]))
gp_offset_bottom_cols[1].metric(tr("recommendation"), target_status)

if offset_result["target_maintained"]:
    st.success(recommendation_message)
else:
    st.error(recommendation_message)

offset_detail_df = pd.DataFrame(offset_result["offset_rows"])
offset_detail_df["reduction_percent"] = offset_detail_df["reduction_percent"] * 100

offset_chart_cols = st.columns(3)
with offset_chart_cols[0]:
    st.subheader(tr("offset_waterfall"))
    offset_waterfall_df = pd.DataFrame(
        {
            "Step": [
                tr("commodity_cost_increase"),
                tr("offset_savings"),
                tr("net_cost_impact"),
            ],
            "Amount": [
                offset_result["commodity_cost_increase"],
                -offset_result["offset_savings"],
                offset_result["net_impact"],
            ],
        }
    ).set_index("Step")
    st.bar_chart(offset_waterfall_df)

with offset_chart_cols[1]:
    st.subheader(tr("offset_category_chart"))
    st.bar_chart(offset_detail_df.set_index("cost_element")["savings"])

with offset_chart_cols[2]:
    st.subheader(tr("gp_before_after_chart"))
    gp_before_after_df = pd.DataFrame(
        {
            "Scenario": [tr("gp_before"), tr("gp_after")],
            "GP": [offset_result["gp_before"], offset_result["gp_after"]],
        }
    ).set_index("Scenario")
    st.bar_chart(gp_before_after_df)

with st.expander(tr("cost_offset_mitigation"), expanded=False):
    st.dataframe(
        offset_detail_df,
        use_container_width=True,
        hide_index=True,
        column_config={
            "category_cost": st.column_config.NumberColumn(format="THB %.0f"),
            "reduction_percent": st.column_config.NumberColumn(format="%.2f%%"),
            "savings": st.column_config.NumberColumn(format="THB %.0f"),
        },
    )

st.subheader(tr("scenario_save_compare"))
scenario_row = {
    "saved_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    "scenario_name": scenario_name.strip(),
    "total_vehicle_cost": round(float(total_new_cost), 2),
    "gp_amount": round(float(total_gp), 2),
    "gp_percent": round(float(total_gp_percent), 4),
    "selected_supplier": selected_supplier,
    "supplier_price_increase_percent": round(float(supplier_price_increase * 100), 4),
    "supplier_delay_days": int(supplier_delay_days),
    "supplier_disruption_percent": round(float(supplier_disruption_rate * 100), 4),
    "delay_penalty_per_day_percent": round(float(delay_penalty_rate * 100), 4),
    "gp_target_percent": round(float(gp_target_percent), 4),
    "selected_module": selected_module,
    "selected_part": selected_part_name,
    "offset_savings": round(float(offset_result["offset_savings"]), 2),
    "offset_net_impact": round(float(offset_result["net_impact"]), 2),
    "offset_gp_after": round(float(offset_result["gp_after"]), 2),
    "offset_gp_after_percent": round(float(offset_result["gp_after_percent"]), 4),
    "offset_required_additional_savings": round(
        float(offset_result["required_additional_savings"]),
        2,
    ),
    "offset_target_maintained": bool(offset_result["target_maintained"]),
}
for driver_name, scenario in price_scenarios.items():
    scenario_row[f"{driver_name} current"] = scenario["current_price"]
    scenario_row[f"{driver_name} scenario"] = scenario["scenario_price"]
    scenario_row[f"{driver_name} change %"] = round(
        float(scenario["change_percent"] * 100),
        4,
    )
for driver_name, config in monte_carlo_config.items():
    scenario_row[f"{driver_name} MC min"] = config["min"]
    scenario_row[f"{driver_name} MC max"] = config["max"]
    scenario_row[f"{driver_name} MC mean"] = config["mean"]
    scenario_row[f"{driver_name} MC volatility"] = config["volatility"]
for category, reduction in offset_reductions.items():
    scenario_row[f"{category} offset reduction %"] = round(float(reduction * 100), 4)

scenario_actions = st.columns([1, 3])
with scenario_actions[0]:
    if st.button(tr("save_scenario"), use_container_width=True):
        if not scenario_row["scenario_name"]:
            st.error(tr("scenario_name_required"))
        else:
            save_scenario(scenario_row)
            st.success(f"{tr('scenario_saved')}: {scenario_row['scenario_name']}")

scenario_history_df = load_scenario_history()
if scenario_history_df.empty:
    st.info(tr("scenario_empty"))
else:
    st.dataframe(
        scenario_history_df.sort_values("saved_at", ascending=False),
        use_container_width=True,
        hide_index=True,
        column_config={
            "total_vehicle_cost": st.column_config.NumberColumn(format="THB %.0f"),
            "gp_amount": st.column_config.NumberColumn(format="THB %.0f"),
            "gp_percent": st.column_config.NumberColumn(format="%.2f%%"),
        },
    )

    scenario_history_df["scenario_label"] = (
        scenario_history_df["saved_at"].astype(str)
        + " | "
        + scenario_history_df["scenario_name"].astype(str)
    )
    compare_cols = st.columns(2)
    first_label = compare_cols[0].selectbox(
        tr("compare_scenario_a"),
        scenario_history_df["scenario_label"].tolist(),
        key="compare_scenario_a",
    )
    second_label = compare_cols[1].selectbox(
        tr("compare_scenario_b"),
        scenario_history_df["scenario_label"].tolist(),
        index=min(1, len(scenario_history_df) - 1),
        key="compare_scenario_b",
    )

    selected_compare = scenario_history_df[
        scenario_history_df["scenario_label"].isin([first_label, second_label])
    ].copy()
    comparison_chart_df = selected_compare.set_index("scenario_label")[
        ["total_vehicle_cost", "gp_amount", "gp_percent"]
    ]
    st.subheader(tr("scenario_comparison_chart"))
    st.bar_chart(comparison_chart_df)

    st.download_button(
        tr("export_excel"),
        data=scenario_history_to_excel(
            scenario_history_df.drop(columns=["scenario_label"])
        ),
        file_name="scenario_history.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )

st.subheader(tr("supplier_risk_dashboard"))
supplier_total_impact = supplier_summary_df["supplier_total_impact"].sum()
supplier_top_cols = st.columns(3, gap="large")
supplier_top_cols[0].metric(tr("active_suppliers"), f"{supplier_summary_df['supplier_name'].nunique():,}")
supplier_top_cols[1].metric(
    tr("supplier_cost_base"),
    format_thb(supplier_summary_df["supplier_cost_base"].sum()),
)
supplier_top_cols[2].metric(tr("supplier_scenario_impact"), format_thb(supplier_total_impact))

supplier_bottom_cols = st.columns(2, gap="large")
supplier_bottom_cols[0].metric(
    tr("fx_impact_by_supplier_country"),
    format_thb(supplier_summary_df["supplier_fx_impact"].sum()),
)
supplier_bottom_cols[1].metric(
    tr("disruption_impact"),
    format_thb(supplier_summary_df["supplier_disruption_impact"].sum()),
)

top_risky_suppliers = supplier_summary_df.sort_values(
    "risk_exposure",
    ascending=False,
)
supplier_dependency = supplier_summary_df.sort_values(
    "vehicle_cost_contribution %",
    ascending=False,
)
supplier_country_heatmap = (
    supplier_summary_df.groupby(["country", "currency"], as_index=False)
    .agg(
        suppliers=("supplier_name", "nunique"),
        supplied_parts=("supplied_parts", "sum"),
        supplier_cost_base=("supplier_cost_base", "sum"),
        supplier_fx_impact=("supplier_fx_impact", "sum"),
        supplier_total_impact=("supplier_total_impact", "sum"),
        risk_exposure=("risk_exposure", "sum"),
    )
    .sort_values("risk_exposure", ascending=False)
)

supplier_panel_cols = st.columns(3)
with supplier_panel_cols[0]:
    st.subheader(tr("top_risky_suppliers"))
    st.dataframe(
        localize_dataframe(top_risky_suppliers[
            [
                "supplier_name",
                "country",
                "currency",
                "risk_level",
                "supplied_parts",
                "supplier_total_impact",
                "risk_exposure",
            ]
        ].head(10)),
        use_container_width=True,
        hide_index=True,
        column_config={
            "supplier_total_impact": st.column_config.NumberColumn(format="THB %.0f"),
            "risk_exposure": st.column_config.NumberColumn(format="THB %.0f"),
        },
    )

with supplier_panel_cols[1]:
    st.subheader(tr("supplier_dependency"))
    st.dataframe(
        localize_dataframe(supplier_dependency[
            [
                "supplier_name",
                "supplied_parts",
                "supplier_cost_base",
                "vehicle_cost_contribution %",
            ]
        ].head(10)),
        use_container_width=True,
        hide_index=True,
        column_config={
            "supplier_cost_base": st.column_config.NumberColumn(format="THB %.0f"),
            "vehicle_cost_contribution %": st.column_config.NumberColumn(format="%.2f%%"),
        },
    )

with supplier_panel_cols[2]:
    st.subheader(tr("supplier_contribution_to_vehicle_cost"))
    st.bar_chart(
        supplier_dependency.head(10).set_index("supplier_name")[
            "vehicle_cost_contribution %"
        ]
    )

st.subheader(tr("supplier_level_cost_impact_analysis"))
st.dataframe(
    localize_dataframe(supplier_summary_df.sort_values(
        "supplier_total_impact",
        key=lambda values: values.abs(),
        ascending=False,
    )),
    use_container_width=True,
    hide_index=True,
    column_config={
        "supplier_cost_base": st.column_config.NumberColumn(format="THB %.0f"),
        "supplier_price_impact": st.column_config.NumberColumn(format="THB %.0f"),
        "supplier_delay_impact": st.column_config.NumberColumn(format="THB %.0f"),
        "supplier_disruption_impact": st.column_config.NumberColumn(format="THB %.0f"),
        "supplier_fx_impact": st.column_config.NumberColumn(format="THB %.0f"),
        "supplier_total_impact": st.column_config.NumberColumn(format="THB %.0f"),
        "vehicle_cost_contribution %": st.column_config.NumberColumn(format="%.2f%%"),
        "risk_exposure": st.column_config.NumberColumn(format="THB %.0f"),
    },
)

st.subheader(tr("supplier_country_heatmap"))
st.dataframe(
    supplier_country_heatmap.style.background_gradient(
        subset=["supplier_total_impact", "risk_exposure"],
        cmap="Reds",
    ),
    use_container_width=True,
)

with st.expander(tr("part_to_supplier_traceability"), expanded=False):
    st.dataframe(
        localize_dataframe(supplier_detail_df[
            [
                "part_code",
                "part_name",
                "module_name",
                "supplier_name",
                "country",
                "currency",
                "risk_level",
                "allocation_percent",
                "supplier_cost_base",
                "supplier_total_impact",
            ]
        ].sort_values("supplier_total_impact", key=lambda values: values.abs(), ascending=False)),
        use_container_width=True,
        hide_index=True,
        column_config={
            "allocation_percent": st.column_config.NumberColumn(format="%.2f"),
            "supplier_cost_base": st.column_config.NumberColumn(format="THB %.0f"),
            "supplier_total_impact": st.column_config.NumberColumn(format="THB %.0f"),
        },
    )

st.subheader(tr("commodity_scenario_exposure"))
driver_impact_rows = []
for driver_name, factor_column in DRIVER_COLUMNS.items():
    affected = parts_df[parts_df[factor_column] > 0]
    affected_modules = ", ".join(sorted(affected["module_name"].unique().tolist()))
    affected_parts = ", ".join(affected["part_name"].head(6).tolist())
    if len(affected) > 6:
        affected_parts = f"{affected_parts} +{len(affected) - 6} more"
    driver_impact_rows.append(
        {
            tr("Driver"): tr(driver_name),
            tr("current"): price_scenarios[driver_name]["current_price"],
            tr("scenario"): price_scenarios[driver_name]["scenario_price"],
            "Unit": price_scenarios[driver_name]["unit"],
            "Calculated impact %": driver_changes[driver_name] * 100,
            "Estimated cost impact": (
                parts_df["total_base_cost"]
                * parts_df[factor_column]
                * driver_changes[driver_name]
            ).sum(),
            "Affected modules": affected_modules,
            "Affected parts": affected_parts,
        }
    )
driver_impact_df = pd.DataFrame(driver_impact_rows).sort_values(
    "Estimated cost impact",
    key=lambda values: values.abs(),
    ascending=False,
)
st.dataframe(
    driver_impact_df,
    use_container_width=True,
    hide_index=True,
    column_config={
        tr("current"): st.column_config.NumberColumn(format="%.2f"),
        tr("scenario"): st.column_config.NumberColumn(format="%.2f"),
        "Calculated impact %": st.column_config.NumberColumn(format="%.2f%%"),
        "Estimated cost impact": st.column_config.NumberColumn(format="THB %.0f"),
    },
)

impact_cols = st.columns(4, gap="large")
largest_move = driver_impact_df.reindex(
    driver_impact_df["Calculated impact %"].abs().sort_values(ascending=False).index
).iloc[0]
largest_exposure = driver_impact_df.iloc[0]
impact_cols[0].metric(
    tr("largest_commodity_move"),
    largest_move[tr("Driver")],
    format_pct(largest_move["Calculated impact %"]),
)
impact_cols[1].metric(
    tr("largest_cost_exposure"),
    largest_exposure[tr("Driver")],
    format_thb(largest_exposure["Estimated cost impact"]),
)
impact_cols[2].metric(
    tr("affected_modules"),
    str(results_df["module_name"].nunique()),
)
impact_cols[3].metric(
    tr("affected_components"),
    str(results_df["part_code"].nunique()),
)

st.subheader(tr("explainability_traceability"))
positive_trace_df = traceability_df[traceability_df["driver_impact"] > 0].copy()
if positive_trace_df.empty:
    positive_trace_df = traceability_df.copy()

driver_contribution = (
    positive_trace_df.groupby("Driver", as_index=False)
    .agg(impact_amount=("driver_impact", "sum"))
    .sort_values("impact_amount", key=lambda values: values.abs(), ascending=False)
)
driver_total_abs_impact = driver_contribution["impact_amount"].abs().sum()
driver_contribution["Contribution %"] = (
    driver_contribution["impact_amount"].abs() / driver_total_abs_impact * 100
    if driver_total_abs_impact
    else 0
)

module_trace_contribution = (
    positive_trace_df.groupby(["Driver", "module_name"], as_index=False)
    .agg(impact_amount=("driver_impact", "sum"))
    .sort_values("impact_amount", key=lambda values: values.abs(), ascending=False)
)
part_trace_contribution = (
    positive_trace_df.groupby(["Driver", "module_name", "part_name"], as_index=False)
    .agg(impact_amount=("driver_impact", "sum"))
    .sort_values("impact_amount", key=lambda values: values.abs(), ascending=False)
)

top_trace_driver = driver_contribution.iloc[0]
top_trace_module = module_trace_contribution[
    module_trace_contribution["Driver"] == top_trace_driver["Driver"]
].iloc[0]
top_trace_part = part_trace_contribution[
    part_trace_contribution["Driver"] == top_trace_driver["Driver"]
].iloc[0]
top_trace_element = positive_trace_df[
    positive_trace_df["Driver"] == top_trace_driver["Driver"]
].sort_values("driver_impact", key=lambda values: values.abs(), ascending=False).iloc[0]
vehicle_cost_direction = "increase" if total_impact >= 0 else "decrease"
base_gp = total_sales_price - total_base_cost
gp_impact = total_gp - base_gp
gp_direction = "decreased" if gp_impact < 0 else "increased"
if total_impact > 0:
    cost_direction = "increased"
elif total_impact < 0:
    cost_direction = "decreased"
else:
    cost_direction = "was unchanged"
top_3_drivers = driver_contribution.head(3).copy()
top_3_modules = (
    positive_trace_df.groupby("module_name", as_index=False)
    .agg(impact_amount=("driver_impact", "sum"))
    .sort_values("impact_amount", key=lambda values: values.abs(), ascending=False)
    .head(3)
)
top_5_parts = (
    positive_trace_df.groupby(["module_name", "part_name"], as_index=False)
    .agg(impact_amount=("driver_impact", "sum"))
    .sort_values("impact_amount", key=lambda values: values.abs(), ascending=False)
    .head(5)
)

st.subheader(tr("executive_explanation_panel"))
if is_base_case:
    executive_message = tr("base_case_no_market_impact")
else:
    executive_message = tr("executive_summary").format(
        vehicle_cost=tr("vehicle_cost"),
        cost_direction=tr(cost_direction),
        impact=format_thb(abs(total_impact)),
        driver=tr(top_trace_driver["Driver"]),
        driver_impact=format_thb(top_trace_driver["impact_amount"]),
        affected_module=tr("affected_module"),
        module=tr(top_trace_module["module_name"]),
        part=top_trace_part["part_name"],
        gross_profit=tr("gross_profit"),
        gp_direction=tr(gp_direction),
        gp_impact=format_thb(abs(gp_impact)),
        gp_percent_label=tr("gp_percent"),
        gp_percent=format_pct(total_gp_percent),
    )
st.info(executive_message)

if not is_base_case:
    st.markdown(
        f"**{tr('why_cost_changed')}:** {tr_root_cause(top_trace_driver['Driver'])}"
    )

executive_cols = st.columns(3)
with executive_cols[0]:
    st.markdown(f"**{tr('top_3_cost_drivers')}**")
    st.dataframe(
        top_3_drivers[["Driver", "impact_amount", "Contribution %"]],
        use_container_width=True,
        hide_index=True,
        column_config={
            "impact_amount": st.column_config.NumberColumn(format="THB %.0f"),
            "Contribution %": st.column_config.NumberColumn(format="%.2f%%"),
        },
    )

with executive_cols[1]:
    st.markdown(f"**{tr('top_3_affected_modules')}**")
    st.dataframe(
        top_3_modules,
        use_container_width=True,
        hide_index=True,
        column_config={
            "impact_amount": st.column_config.NumberColumn(format="THB %.0f"),
        },
    )

with executive_cols[2]:
    st.markdown(f"**{tr('top_5_affected_parts')}**")
    st.dataframe(
        top_5_parts,
        use_container_width=True,
        hide_index=True,
        column_config={
            "impact_amount": st.column_config.NumberColumn(format="THB %.0f"),
        },
    )

chain_cols = st.columns(4)
chain_cols[0].metric(
    tr("root_driver"),
    f"{tr(top_trace_driver['Driver'])} {tr('up')}",
    format_pct(driver_changes[top_trace_driver["Driver"]] * 100),
)
chain_cols[1].metric(tr("affected_part"), top_trace_part["part_name"])
chain_cols[2].metric(tr("affected_module"), tr(top_trace_module["module_name"]))
chain_cols[3].metric(tr("vehicle_cost"), tr(vehicle_cost_direction), format_thb(total_impact))

st.markdown(
    f"**{tr('impact_chain')}:** {tr(top_trace_driver['Driver'])} {tr('up')} -> "
    f"{top_trace_part['part_name']} {tr('up')} -> {tr(top_trace_module['module_name'])} {tr('up')} -> "
    f"{tr('vehicle_cost')} {tr(vehicle_cost_direction)}"
)

trace_cols = st.columns(3)
with trace_cols[0]:
    st.subheader(tr("contribution_by_driver"))
    st.dataframe(
        localize_dataframe(driver_contribution),
        use_container_width=True,
        hide_index=True,
        column_config={
            "impact_amount": st.column_config.NumberColumn(format="THB %.0f"),
            "Contribution %": st.column_config.NumberColumn(format="%.2f%%"),
        },
    )

with trace_cols[1]:
    st.subheader(tr("contribution_by_module"))
    st.dataframe(
        localize_dataframe(module_trace_contribution.head(10)),
        use_container_width=True,
        hide_index=True,
        column_config={
            "impact_amount": st.column_config.NumberColumn(format="THB %.0f"),
        },
    )

with trace_cols[2]:
    st.subheader(tr("contribution_by_part"))
    st.dataframe(
        localize_dataframe(part_trace_contribution.head(10)),
        use_container_width=True,
        hide_index=True,
        column_config={
            "impact_amount": st.column_config.NumberColumn(format="THB %.0f"),
        },
    )

waterfall_trace_df = pd.concat(
    [
        pd.DataFrame({"Step": ["Base vehicle cost"], "Impact": [total_base_cost]}),
        driver_contribution.rename(
            columns={"Driver": "Step", "impact_amount": "Impact"}
        )[["Step", "Impact"]],
        pd.DataFrame({"Step": ["Simulated vehicle cost"], "Impact": [total_new_cost]}),
    ],
    ignore_index=True,
)
st.subheader(tr("waterfall_impact_trace"))
st.bar_chart(waterfall_trace_df.set_index("Step")["Impact"])

st.subheader(tr("root_cause_trace_table"))
trace_table = positive_trace_df[
    [
        "Driver",
        "Root cause",
        "module_name",
        "part_name",
        "cost_element",
        "driver_impact",
        "Explanation",
    ]
].sort_values("driver_impact", key=lambda values: values.abs(), ascending=False)
st.dataframe(
    localize_dataframe(trace_table.head(25)),
    use_container_width=True,
    hide_index=True,
    column_config={
        "driver_impact": st.column_config.NumberColumn(format="THB %.0f"),
    },
)

st.subheader(tr("sensitivity_analysis"))
sensitivity_rows = []
sensitivity_part_frames = []
for driver_name in DRIVER_COLUMNS:
    isolated_changes = {name: 0.0 for name in DRIVER_COLUMNS}
    isolated_changes[driver_name] = driver_changes[driver_name]
    driver_breakdown = calculate_element_impacts(breakdown_enriched, isolated_changes)
    driver_parts = (
        driver_breakdown.groupby(
            ["part_code", "part_name", "module_name"],
            as_index=False,
        )
        .agg(driver_impact=("element_impact", "sum"))
    )
    driver_parts["driver"] = driver_name
    sensitivity_part_frames.append(driver_parts)

    cost_impact = driver_parts["driver_impact"].sum()
    gp_impact = -cost_impact
    gp_percent_impact = (
        gp_impact / total_sales_price * 100
        if total_sales_price
        else 0
    )
    sensitivity_rows.append(
        {
            "Driver": driver_name,
            "Calculated impact %": driver_changes[driver_name] * 100,
            "Impact on total vehicle cost": cost_impact,
            "Impact on GP": gp_impact,
            "Impact on GP%": gp_percent_impact,
            "Risk score": abs(gp_impact),
        }
    )

sensitivity_df = pd.DataFrame(sensitivity_rows)
sensitivity_df["Risk level"] = pd.qcut(
    sensitivity_df["Risk score"].rank(method="first"),
    q=4,
    labels=["Low", "Medium", "High", "Critical"],
)
profitability_ranking = sensitivity_df.sort_values(
    "Impact on GP",
    key=lambda values: values.abs(),
    ascending=False,
)

tornado_df = profitability_ranking[["Driver", "Impact on GP"]].set_index("Driver")

sensitivity_cols = st.columns(2)
with sensitivity_cols[0]:
    st.subheader(tr("top_drivers_profitability"))
    st.dataframe(
        localize_dataframe(profitability_ranking[
            [
                "Driver",
                "Calculated impact %",
                "Impact on total vehicle cost",
                "Impact on GP",
                "Impact on GP%",
                "Risk level",
            ]
        ]),
        use_container_width=True,
        hide_index=True,
        column_config={
            "Calculated impact %": st.column_config.NumberColumn(format="%.2f%%"),
            "Impact on total vehicle cost": st.column_config.NumberColumn(format="THB %.0f"),
            "Impact on GP": st.column_config.NumberColumn(format="THB %.0f"),
            "Impact on GP%": st.column_config.NumberColumn(format="%.2f pts"),
        },
    )

with sensitivity_cols[1]:
    st.subheader(tr("tornado_chart"))
    st.bar_chart(tornado_df)

all_driver_parts = pd.concat(sensitivity_part_frames, ignore_index=True)
top_driver = profitability_ranking.iloc[0]["Driver"]
top_driver_parts = all_driver_parts[all_driver_parts["driver"] == top_driver].copy()
module_contribution = (
    top_driver_parts.groupby("module_name", as_index=False)
    .agg(impact_amount=("driver_impact", "sum"))
    .sort_values("impact_amount", key=lambda values: values.abs(), ascending=False)
)
part_contribution = top_driver_parts.sort_values(
    "driver_impact",
    key=lambda values: values.abs(),
    ascending=False,
).head(10)

contribution_cols = st.columns(2)
with contribution_cols[0]:
    st.subheader(tr("most_sensitive_module"))
    st.dataframe(localize_dataframe(module_contribution), use_container_width=True, hide_index=True)

with contribution_cols[1]:
    st.subheader(tr("most_sensitive_part"))
    st.dataframe(localize_dataframe(part_contribution), use_container_width=True, hide_index=True)

top_module = module_contribution.iloc[0]
top_part = part_contribution.iloc[0]
total_top_driver_impact = top_driver_parts["driver_impact"].sum()
module_share = (
    abs(top_module["impact_amount"]) / abs(total_top_driver_impact) * 100
    if total_top_driver_impact
    else 0
)
impact_direction = "increase" if total_top_driver_impact >= 0 else "decrease"
st.info(
    f"{top_module['module_name']} contributes {module_share:.0f}% of total cost {impact_direction} "
    f"due to {top_driver} movement. The most sensitive component is "
    f"{top_part['part_name']} with {format_thb(top_part['driver_impact'])} impact."
)

st.subheader(tr("monte_carlo_simulation"))
rng = np.random.default_rng(42)
driver_names = list(DRIVER_COLUMNS)
sampled_driver_values = {}
sampled_driver_changes = []
for driver_name in driver_names:
    config = monte_carlo_config[driver_name]
    current_price = price_scenarios[driver_name]["current_price"]
    sampled_values = rng.normal(
        loc=config["mean"],
        scale=config["volatility"],
        size=MONTE_CARLO_RUNS,
    )
    sampled_values = np.clip(sampled_values, config["min"], config["max"])
    sampled_driver_values[driver_name] = sampled_values
    sampled_driver_changes.append((sampled_values - current_price) / current_price)

driver_change_matrix = np.column_stack(sampled_driver_changes)
row_driver_weights = build_row_driver_weights(breakdown_enriched)
row_change_matrix = driver_change_matrix @ row_driver_weights.T
base_cost_vector = breakdown_enriched["cost_amount"].to_numpy()
simulation_cost_impact = row_change_matrix * base_cost_vector
total_cost_distribution = total_base_cost + simulation_cost_impact.sum(axis=1)
gp_distribution = total_sales_price - total_cost_distribution
gp_percent_distribution = gp_distribution / total_sales_price * 100

gp_below_target_probability = float(
    (gp_percent_distribution < gp_target_percent).mean() * 100
)
worst_index = int(np.argmin(gp_percent_distribution))
best_index = int(np.argmax(gp_percent_distribution))
average_cost = float(total_cost_distribution.mean())
average_gp = float(gp_distribution.mean())
average_gp_percent = float(gp_percent_distribution.mean())
var_gp_95 = float(np.percentile(gp_distribution, 5))
var_cost_95 = float(np.percentile(total_cost_distribution, 95))
mc_risk_level = classify_monte_carlo_risk(gp_below_target_probability / 100)

mc_top_cols = st.columns(3, gap="large")
mc_top_cols[0].metric(tr("simulations"), f"{MONTE_CARLO_RUNS:,}")
mc_top_cols[1].metric(tr("probability_gp_below_target"), format_pct(gp_below_target_probability))
mc_top_cols[2].metric(tr("average_gp_percent"), format_pct(average_gp_percent))

mc_bottom_cols = st.columns(2, gap="large")
mc_bottom_cols[0].metric(tr("value_at_risk"), format_thb(var_gp_95))
mc_bottom_cols[1].metric(tr("risk_classification"), tr(mc_risk_level.lower().replace(" ", "_")))

case_df = pd.DataFrame(
    [
        {
            "Case": "Worst case",
            "Total vehicle cost": total_cost_distribution[worst_index],
            "GP": gp_distribution[worst_index],
            "GP%": gp_percent_distribution[worst_index],
        },
        {
            "Case": "Average case",
            "Total vehicle cost": average_cost,
            "GP": average_gp,
            "GP%": average_gp_percent,
        },
        {
            "Case": "Best case",
            "Total vehicle cost": total_cost_distribution[best_index],
            "GP": gp_distribution[best_index],
            "GP%": gp_percent_distribution[best_index],
        },
        {
            "Case": "VaR 95",
            "Total vehicle cost": var_cost_95,
            "GP": var_gp_95,
            "GP%": np.percentile(gp_percent_distribution, 5),
        },
    ]
)
st.dataframe(
    case_df,
    use_container_width=True,
    hide_index=True,
    column_config={
        "Total vehicle cost": st.column_config.NumberColumn(format="THB %.0f"),
        "GP": st.column_config.NumberColumn(format="THB %.0f"),
        "GP%": st.column_config.NumberColumn(format="%.2f%%"),
    },
)

st.info(
    f"There is a {gp_below_target_probability:.0f}% probability that GP% will fall below "
    f"{gp_target_percent:.1f}% under current commodity volatility. The simulated portfolio is "
    f"classified as {mc_risk_level}."
)

hist_counts, hist_bins = np.histogram(total_cost_distribution, bins=40)
histogram_df = pd.DataFrame(
    {
        "Total vehicle cost": hist_bins[:-1],
        "Frequency": hist_counts,
    }
).set_index("Total vehicle cost")

density_counts, density_bins = np.histogram(gp_percent_distribution, bins=40, density=True)
density_df = pd.DataFrame(
    {
        "GP%": density_bins[:-1],
        "Density": density_counts,
    }
).set_index("GP%")

probability_df = pd.DataFrame(
    {
        "Outcome": ["GP% below target", "GP% at or above target"],
        "Probability": [
            gp_below_target_probability,
            100 - gp_below_target_probability,
        ],
    }
).set_index("Outcome")

module_index = breakdown_enriched["module_name"].astype("category")
module_names = list(module_index.cat.categories)
module_codes = module_index.cat.codes.to_numpy()
module_impacts = pd.DataFrame(index=driver_names, columns=module_names, dtype=float)
for driver_position, driver_name in enumerate(driver_names):
    driver_only_change = driver_change_matrix[:, driver_position]
    driver_row_weights = row_driver_weights[:, driver_position]
    driver_row_impact = (
        driver_only_change[:, None]
        * driver_row_weights[None, :]
        * base_cost_vector[None, :]
    )
    for module_position, module_name in enumerate(module_names):
        module_mask = module_codes == module_position
        module_impacts.loc[driver_name, module_name] = float(
            np.mean(np.abs(driver_row_impact[:, module_mask].sum(axis=1)))
        )

mc_chart_cols = st.columns(2)
with mc_chart_cols[0]:
    st.subheader(tr("histogram_total_vehicle_cost"))
    st.bar_chart(histogram_df)

with mc_chart_cols[1]:
    st.subheader(tr("distribution_curve_gp"))
    st.line_chart(density_df)

mc_chart_cols = st.columns(2)
with mc_chart_cols[0]:
    st.subheader(tr("risk_heatmap"))
    st.dataframe(
        module_impacts.style.background_gradient(axis=None, cmap="Reds"),
        use_container_width=True,
    )

with mc_chart_cols[1]:
    st.subheader(tr("probability_chart"))
    st.bar_chart(probability_df)

module_summary = (
    results_df.groupby("module_name", as_index=False)
    .agg(
        components=("part_code", "count"),
        sales_price=("sales_price", "sum"),
        old_cost=("old_cost", "sum"),
        new_cost=("new_cost", "sum"),
        impact_amount=("impact_amount", "sum"),
        gp_amount=("gp_amount", "sum"),
        risk_exposure=("risk_exposure", "sum"),
    )
    .sort_values("new_cost", ascending=False)
)
module_summary["gp_percent"] = module_summary["gp_amount"] / module_summary["sales_price"] * 100

st.subheader(tr("module_level_summary"))
st.dataframe(localize_dataframe(module_summary), use_container_width=True, hide_index=True)

st.subheader(tr("hierarchical_bom_tree"))
for module_name, module_df in results_df.groupby("module_name", sort=True):
    module_cost = module_df["new_cost"].sum()
    module_risk = module_df["risk_exposure"].sum()
    with st.expander(
        f"{tr(module_name)} | {len(module_df)} {tr('component')} | {format_thb(module_cost)} | {tr('risk_level')} {format_thb(module_risk)}",
        expanded=module_name == selected_part["module_name"],
    ):
        for _, parent in module_df[module_df["bom_level"] == 1].iterrows():
            st.markdown(f"**{parent['part_code']} - {parent['part_name']}**")
            children = module_df[module_df["parent_part_code"] == parent["part_code"]]
            tree_rows = pd.concat([parent.to_frame().T, children], ignore_index=True)
            st.dataframe(
                localize_dataframe(tree_rows[
                    [
                        "bom_level",
                        "component_type",
                        "part_code",
                        "part_name",
                        "material_group",
                        "old_cost",
                        "new_cost",
                        "impact_amount",
                        "risk_score",
                    ]
                ]),
                use_container_width=True,
                hide_index=True,
            )

chart_cols = st.columns(2)
with chart_cols[0]:
    st.subheader(tr("cost_by_module"))
    st.bar_chart(module_summary.set_index("module_name")[["old_cost", "new_cost"]])

with chart_cols[1]:
    st.subheader(tr("top_impacted_components"))
    top_impacted = results_df.reindex(
        results_df["impact_amount"].abs().sort_values(ascending=False).index
    ).head(10)
    st.bar_chart(top_impacted.set_index("part_name")["impact_amount"])

chart_cols = st.columns(2)
with chart_cols[0]:
    st.subheader(tr("top_10_expensive_components"))
    top_expensive = results_df.nlargest(10, "new_cost").set_index("part_name")["new_cost"]
    st.bar_chart(top_expensive)

with chart_cols[1]:
    st.subheader(tr("risk_analysis"))
    risk_chart = results_df.nlargest(10, "risk_exposure").set_index("part_name")[
        "risk_exposure"
    ]
    st.bar_chart(risk_chart)

st.subheader(tr("waterfall_cost_analysis"))
waterfall_df = pd.DataFrame(
    [
        {"step": "Base cost", "amount": total_base_cost},
        {"step": "Driver impact", "amount": total_impact},
        {"step": "Simulated cost", "amount": total_new_cost},
        {"step": "Gross profit", "amount": total_gp},
    ]
)
st.bar_chart(waterfall_df.set_index("step")["amount"])

st.subheader(tr("part_level_summary"))
part_table = filtered_df[
    [
        "part_code",
        "parent_part_code",
        "bom_level",
        "component_type",
        "part_name",
        "module_name",
        "material_group",
        "sales_price",
        "old_cost",
        "new_cost",
        "impact_amount",
        "gp_amount",
        "gp_percent",
        "risk_score",
        "risk_exposure",
        "risk_band",
    ]
].copy()
part_table["gp_percent"] = part_table["gp_percent"].map(format_pct)
st.dataframe(localize_dataframe(part_table), use_container_width=True, hide_index=True)

st.subheader(tr("selected_component_breakdown"))
selected_top_cols = st.columns(3, gap="large")
selected_top_cols[0].metric(tr("component"), selected_part["part_code"])
selected_top_cols[1].metric(tr("base_cost"), format_thb(selected_part["old_cost"]))
selected_top_cols[2].metric(tr("simulated_cost"), format_thb(selected_part["new_cost"]))

selected_bottom_cols = st.columns(2, gap="large")
selected_bottom_cols[0].metric(tr("impact"), format_thb(selected_part["impact_amount"]))
selected_bottom_cols[1].metric(tr("risk_level"), tr(str(selected_part["risk_band"]).lower()))

breakdown_cols = st.columns(2)
with breakdown_cols[0]:
    breakdown_table = selected_breakdown[
        ["cost_element", "cost_amount", "element_impact", "new_cost_amount"]
    ].copy()
    st.dataframe(localize_dataframe(breakdown_table), use_container_width=True, hide_index=True)

with breakdown_cols[1]:
    st.subheader(tr("selected_component_cost_breakdown"))
    st.bar_chart(selected_breakdown.set_index("cost_element")["new_cost_amount"])

st.subheader(tr("before_vs_after_simulation"))
comparison_df = pd.DataFrame(
    {"Base cost": results_df["old_cost"], "New cost": results_df["new_cost"]},
    index=results_df["part_name"],
)
st.bar_chart(comparison_df)
